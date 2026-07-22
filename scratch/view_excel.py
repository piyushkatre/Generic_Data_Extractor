import openpyxl

wb = openpyxl.load_workbook("datasets/franchise_bazar.xlsx")
sheet = wb.active

print(f"Sheet Title: {sheet.title}")
print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")

# Print headers
headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
print("\nHeaders:")
print(headers)

# Print row 2 values
if sheet.max_row >= 2:
    print("\nRow 2 values:")
    for col_idx in range(1, sheet.max_column + 1):
        header = headers[col_idx - 1]
        val = sheet.cell(row=2, column=col_idx).value
        print(f"  {header}: {val}")
else:
    print("No data rows found in the sheet.")

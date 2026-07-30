"""
Tests for the format-agnostic dataset output layer (writers/).

Covers:
- writers.dataset_writer.create_writer() format selection + DatasetBuilder
  dependency injection.
- CSVDatasetWriter spec compliance (UTF-8, newline="", header-once,
  column-order preservation, blank-for-missing, deterministic output,
  append-only/no-dedup - a documented behavior difference from Excel).
- Regression: the same schema's columns appear in identical order in an
  Excel workbook and a CSV file (Step 7 of the Excel->CSV migration).
- DatasetBuilder.save_extraction_result() end to end with output_format="csv".
"""

import csv
import os

import openpyxl
import pytest

from modules.dataset_builder.builder import DatasetBuilder
from writers.dataset_writer import create_writer
from writers.excel_writer import ExcelDatasetWriter
from writers.csv_writer import CSVDatasetWriter


SCHEMA = {
    "dataset_name": "widgets.xlsx",
    "sheet_name": "Widgets",
    "primary_key": ["Source URL"],
    "columns": ["Source URL", "Franchise Name", "Phone", "Additional Information", "Extraction Date", "Last Updated"],
}


# ---------------------------------------------------------------------
# Writer selection / dependency injection
# ---------------------------------------------------------------------

def test_create_writer_factory_selects_by_output_format(tmp_path):
    excel_writer = create_writer("excel", datasets_dir=str(tmp_path))
    assert isinstance(excel_writer, ExcelDatasetWriter)

    csv_writer = create_writer("csv", datasets_dir=str(tmp_path))
    assert isinstance(csv_writer, CSVDatasetWriter)

    # Case/whitespace tolerant, since it's meant to come straight from a
    # config value.
    assert isinstance(create_writer(" CSV ", datasets_dir=str(tmp_path)), CSVDatasetWriter)

    with pytest.raises(ValueError):
        create_writer("json", datasets_dir=str(tmp_path))


def test_dataset_builder_defaults_to_excel_writer(tmp_path):
    builder = DatasetBuilder(datasets_dir=str(tmp_path))
    assert isinstance(builder.writer, ExcelDatasetWriter)


def test_dataset_builder_output_format_selects_csv_writer(tmp_path):
    builder = DatasetBuilder(datasets_dir=str(tmp_path), output_format="csv")
    assert isinstance(builder.writer, CSVDatasetWriter)


def test_dataset_builder_dependency_injection_uses_given_writer(tmp_path):
    """Prefer dependency injection: an explicit writer= bypasses the
    output_format lookup entirely."""
    writer = CSVDatasetWriter(datasets_dir=str(tmp_path))
    builder = DatasetBuilder(datasets_dir=str(tmp_path), output_format="excel", writer=writer)
    assert builder.writer is writer


# ---------------------------------------------------------------------
# CSVDatasetWriter spec compliance
# ---------------------------------------------------------------------

def test_csv_writer_basic_spec_compliance(tmp_path):
    writer = CSVDatasetWriter(datasets_dir=str(tmp_path))
    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    writer.write_headers()

    op1 = writer.append_record(
        {"Source URL": "https://a.com", "Franchise Name": "Alpha Gym", "Phone": "111"},
        primary_keys=["Source URL"], timestamp="2026-01-01 00:00:00",
    )
    op2 = writer.append_record(
        {"Source URL": "https://b.com", "Franchise Name": "Beta Gym"},  # Phone missing
        primary_keys=["Source URL"], timestamp="2026-01-01 00:05:00",
    )
    info = writer.finish()

    assert op1 == {"operation": "Inserted", "row_number": 1}
    assert op2 == {"operation": "Inserted", "row_number": 2}

    file_path = info["file_path"]
    assert file_path.endswith(".csv")

    with open(file_path, "r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))

    assert rows[0] == SCHEMA["columns"]  # header written once, column order preserved
    assert rows[1][0] == "https://a.com"
    assert rows[1][1] == "Alpha Gym"
    assert rows[1][2] == "111"
    assert rows[2][0] == "https://b.com"
    assert rows[2][2] == ""  # blank string for missing value, not "None"/"null"

    # Extraction Date / Last Updated set on every insert, mirroring
    # ExcelDatasetWriter's insert path.
    ed_idx = SCHEMA["columns"].index("Extraction Date")
    lu_idx = SCHEMA["columns"].index("Last Updated")
    assert rows[1][ed_idx] == "2026-01-01 00:00:00"
    assert rows[1][lu_idx] == "2026-01-01 00:00:00"


def test_csv_writer_header_written_once_across_repeated_begin_dataset_calls(tmp_path):
    """DatasetBuilder calls begin_dataset()/write_headers() again for every
    URL against the same dataset - the header must never repeat."""
    writer = CSVDatasetWriter(datasets_dir=str(tmp_path))

    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    writer.write_headers()
    writer.append_record({"Source URL": "https://a.com", "Franchise Name": "Alpha"}, ["Source URL"], "2026-01-01 00:00:00")
    writer.finish()

    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    writer.write_headers()
    writer.append_record({"Source URL": "https://b.com", "Franchise Name": "Beta"}, ["Source URL"], "2026-01-01 00:05:00")
    info = writer.finish()

    with open(info["file_path"], "r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))

    assert rows[0] == SCHEMA["columns"]
    assert len(rows) == 3  # header + 2 data rows, no duplicate header
    assert [r[1] for r in rows[1:]] == ["Alpha", "Beta"]


def test_csv_writer_is_append_only_no_dedup(tmp_path):
    """Documented behavior difference from Excel: CSV has no efficient
    random-access update, so it always inserts, even for a repeated primary
    key - there is no update-in-place / merge."""
    writer = CSVDatasetWriter(datasets_dir=str(tmp_path))
    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    writer.write_headers()
    writer.append_record({"Source URL": "https://a.com", "Franchise Name": "Alpha"}, ["Source URL"], "2026-01-01 00:00:00")
    op2 = writer.append_record({"Source URL": "https://a.com", "Franchise Name": "Alpha v2"}, ["Source URL"], "2026-01-01 01:00:00")
    info = writer.finish()

    assert op2["operation"] == "Inserted"  # never "Updated"
    with open(info["file_path"], "r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))
    assert len(rows) == 3  # header + both rows kept, no merge


def test_csv_writer_utf8_encoding_roundtrip(tmp_path):
    writer = CSVDatasetWriter(datasets_dir=str(tmp_path))
    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], ["Source URL", "Investment Required"])
    writer.write_headers()
    writer.append_record(
        {"Source URL": "https://a.com", "Investment Required": "₹15 Lakhs - ₹20 Lakhs"},
        primary_keys=["Source URL"], timestamp="2026-01-01 00:00:00",
    )
    info = writer.finish()

    with open(info["file_path"], "r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))
    assert rows[1][1] == "₹15 Lakhs - ₹20 Lakhs"

    # Confirm the bytes on disk are genuinely UTF-8, not e.g. cp1252 (which
    # would encode/represent the rupee sign differently, or fail outright).
    with open(info["file_path"], "rb") as f:
        raw = f.read()
    assert "₹".encode("utf-8") in raw


def test_csv_writer_handles_embedded_commas_and_newlines(tmp_path):
    """Validates the newline="" + csv.writer contract: an embedded newline
    inside a quoted field must not be corrupted into a stray \\r\\r\\n or
    split into an extra row (the classic bug from opening a CSV in text
    mode on Windows without newline="")."""
    writer = CSVDatasetWriter(datasets_dir=str(tmp_path))
    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], ["Source URL", "Additional Information"])
    writer.write_headers()
    tricky_value = "Line one\nLine two, with a comma"
    writer.append_record(
        {"Source URL": "https://a.com", "Additional Information": tricky_value},
        primary_keys=["Source URL"], timestamp="2026-01-01 00:00:00",
    )
    info = writer.finish()

    with open(info["file_path"], "r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))
    assert len(rows) == 2  # header + exactly one data row
    assert rows[1][1] == tricky_value


def test_csv_writer_deterministic_output_is_byte_identical_across_runs(tmp_path):
    def run(dir_path):
        writer = CSVDatasetWriter(datasets_dir=dir_path)
        writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
        writer.write_headers()
        writer.append_record({"Source URL": "https://a.com", "Franchise Name": "Alpha"}, ["Source URL"], "2026-01-01 00:00:00")
        info = writer.finish()
        with open(info["file_path"], "rb") as f:
            return f.read()

    dir1 = tmp_path / "run1"
    dir2 = tmp_path / "run2"
    dir1.mkdir()
    dir2.mkdir()
    assert run(str(dir1)) == run(str(dir2))


# ---------------------------------------------------------------------
# Step 7 regression: Excel columns vs CSV headers for the same schema
# ---------------------------------------------------------------------

def test_excel_and_csv_writers_produce_identical_column_order(tmp_path):
    excel_dir = tmp_path / "excel_out"
    csv_dir = tmp_path / "csv_out"
    excel_dir.mkdir()
    csv_dir.mkdir()

    record = {
        "Source URL": "https://cultfit.com",
        "Franchise Name": "Cult Fit",
        "Phone": "1234567890",
    }

    excel_writer = ExcelDatasetWriter(datasets_dir=str(excel_dir))
    excel_writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    excel_writer.write_headers()
    excel_writer.append_record(dict(record), primary_keys=SCHEMA["primary_key"], timestamp="2026-01-01 00:00:00")
    excel_info = excel_writer.finish()

    csv_writer = CSVDatasetWriter(datasets_dir=str(csv_dir))
    csv_writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    csv_writer.write_headers()
    csv_writer.append_record(dict(record), primary_keys=SCHEMA["primary_key"], timestamp="2026-01-01 00:00:00")
    csv_info = csv_writer.finish()

    wb = openpyxl.load_workbook(excel_info["file_path"])
    ws = wb[SCHEMA["sheet_name"]]
    excel_headers = [cell.value for cell in ws[1]]

    with open(csv_info["file_path"], "r", encoding="utf-8", newline="") as f:
        csv_headers = next(csv.reader(f))

    assert excel_headers == SCHEMA["columns"]
    assert csv_headers == SCHEMA["columns"]
    assert excel_headers == csv_headers  # every mapped field, identical order, both formats

    excel_row = [cell.value for cell in ws[2]]
    excel_row_dict = dict(zip(excel_headers, excel_row))
    with open(csv_info["file_path"], "r", encoding="utf-8", newline="") as f:
        csv_row = list(csv.reader(f))[1]
    csv_row_dict = dict(zip(csv_headers, csv_row))

    assert excel_row_dict["Franchise Name"] == csv_row_dict["Franchise Name"] == "Cult Fit"
    assert excel_row_dict["Phone"] == csv_row_dict["Phone"] == "1234567890"
    assert (excel_row_dict["Source URL"] or "") == (csv_row_dict["Source URL"] or "") == "https://cultfit.com"


# ---------------------------------------------------------------------
# DatasetBuilder end to end with output_format="csv"
# ---------------------------------------------------------------------

def test_dataset_builder_csv_output_end_to_end(tmp_path):
    """Same DatasetBuilder.save_extraction_result() entry point used by
    core/pipeline.py, just with output_format="csv" - verifies the whole
    normalized-record -> RecordMapper -> writer path, not only the writer
    classes in isolation."""
    from modules.gemini import ExtractionResult

    schema = {
        "dataset_name": "csv_e2e_test.xlsx",
        "sheet_name": "Franchise Data",
        "primary_key": ["Source URL", "Franchise Name"],
        "columns": ["Source URL", "Franchise Name", "Investment Required", "Extraction Date", "Last Updated"],
        # Not under test here - see tests/test_metadata_columns.py for
        # DatasetBuilder's metadata_columns behavior.
        "metadata_columns": [],
    }
    result = ExtractionResult(
        franchise_name="CSV Test Gym",
        investment_required="Rs. 10 Lakhs",
        page_type="Franchise Listing",
    )

    builder = DatasetBuilder(datasets_dir=str(tmp_path), output_format="csv")
    save_info = builder.save_extraction_result(
        result=result,
        source_url="https://csvtest.com",
        detected_page_type="Franchise Listing",
        timestamp="2026-01-01 00:00:00",
        schema=schema,
        page_type="Franchise Listing",
    )

    assert save_info["status"] == "Success"
    assert save_info["operation"] == "Inserted"
    assert save_info["workbook_name"].endswith(".csv")

    csv_path = os.path.join(str(tmp_path), "csv_e2e_test.csv")
    assert os.path.exists(csv_path)
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))
    assert rows[0] == schema["columns"]
    row_dict = dict(zip(rows[0], rows[1]))
    assert row_dict["Franchise Name"] == "CSV Test Gym"
    assert "10 Lakhs" in row_dict["Investment Required"]

    # UI-facing fields (app/ui/_dataset_preview.py): the real, directly
    # openable path (not the legacy "datasets/"-prefixed workbook_name),
    # the format, and the row count - so the UI never has to search
    # datasets_dir or guess an extension.
    assert save_info["output_path"] == csv_path
    assert os.path.exists(save_info["output_path"])
    assert save_info["output_format"] == "csv"
    assert save_info["records_written"] == 1


# ---------------------------------------------------------------------
# writers.DatasetWriter.finish() contract: format/record_count/sheet_name
# (app/ui/_dataset_preview.py's only source of "which file, which format,
# how many rows, which sheet" - never re-derived by scanning a directory)
# ---------------------------------------------------------------------

def test_csv_writer_finish_reports_format_and_record_count(tmp_path):
    writer = CSVDatasetWriter(datasets_dir=str(tmp_path))
    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    writer.write_headers()
    writer.append_record({"Source URL": "https://a.com", "Franchise Name": "Alpha"}, ["Source URL"], "2026-01-01 00:00:00")
    writer.append_record({"Source URL": "https://b.com", "Franchise Name": "Beta"}, ["Source URL"], "2026-01-01 00:05:00")
    info = writer.finish()

    assert info["format"] == "csv"
    assert info["record_count"] == 2
    assert info["file_path"].endswith(".csv")
    assert os.path.exists(info["file_path"])


def test_csv_writer_finish_record_count_reflects_whole_file_not_just_this_call(tmp_path):
    """A second DatasetBuilder call against the same dataset (the second
    URL in a batch job) must report the TOTAL row count in the file, not
    just what this call itself wrote - this is what makes
    get_latest_save_info() correct for batch jobs without any special-case
    aggregation in the UI."""
    writer = CSVDatasetWriter(datasets_dir=str(tmp_path))

    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    writer.write_headers()
    writer.append_record({"Source URL": "https://a.com", "Franchise Name": "Alpha"}, ["Source URL"], "2026-01-01 00:00:00")
    first_info = writer.finish()
    assert first_info["record_count"] == 1

    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    writer.write_headers()
    writer.append_record({"Source URL": "https://b.com", "Franchise Name": "Beta"}, ["Source URL"], "2026-01-01 00:05:00")
    second_info = writer.finish()
    assert second_info["record_count"] == 2


def test_excel_writer_finish_reports_format_record_count_and_sheet_name(tmp_path):
    writer = ExcelDatasetWriter(datasets_dir=str(tmp_path))
    writer.begin_dataset(SCHEMA["dataset_name"], SCHEMA["sheet_name"], SCHEMA["columns"])
    writer.write_headers()
    writer.append_record({"Source URL": "https://a.com", "Franchise Name": "Alpha"}, ["Source URL"], "2026-01-01 00:00:00")
    info = writer.finish()

    assert info["format"] == "excel"
    assert info["record_count"] == 1
    assert info["sheet_name"] == SCHEMA["sheet_name"]
    assert os.path.exists(info["file_path"])

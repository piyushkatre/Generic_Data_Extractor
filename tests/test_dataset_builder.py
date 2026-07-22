import os
import json
import shutil
import pytest
import openpyxl
from unittest.mock import patch, MagicMock

from modules.dataset_builder.schema_loader import SchemaLoader
from modules.dataset_builder.normalizer import RecordNormalizer
from modules.dataset_builder.detector import DuplicateDetector
from modules.dataset_builder.manager import WorkbookManager, ExcelWriter
from modules.dataset_builder.builder import DatasetBuilder
from modules.gemini import ExtractionResult, ExtractedEntity, ExtractedRecord, KeyValue

TEST_SCHEMAS_DIR = "test_schemas"
TEST_DATASETS_DIR = "test_datasets"

@pytest.fixture(scope="module", autouse=True)
def setup_test_directories():
    # Setup test directories
    os.makedirs(TEST_SCHEMAS_DIR, exist_ok=True)
    os.makedirs(TEST_DATASETS_DIR, exist_ok=True)
    
    # Create a couple of mock schemas for testing
    franchise_mock = {
        "dataset_name": "test_franchise.xlsx",
        "sheet_name": "Franchise Data",
        "primary_key": ["Source URL", "Franchise Name"],
        "required_fields": ["Franchise Name", "Source URL"],
        "aliases": {
            "franchise name": "Franchise Name",
            "source url": "Source URL",
            "investment": "Investment Required",
            "phone": "Phone",
            "website": "Website"
        },
        "columns": ["Source URL", "Franchise Name", "Brand", "Investment Required", "Phone", "Website", "Additional Information", "Extraction Date", "Last Updated"]
    }
    
    misc_mock = {
        "dataset_name": "test_misc.xlsx",
        "sheet_name": "General Data",
        "primary_key": ["Source URL", "Title"],
        "required_fields": ["Title", "Source URL"],
        "aliases": {
            "title": "Title",
            "source url": "Source URL"
        },
        "columns": ["Source URL", "Title", "Additional Information", "Extraction Date", "Last Updated"]
    }
    
    with open(os.path.join(TEST_SCHEMAS_DIR, "franchise_schema.json"), "w", encoding="utf-8") as f:
        json.dump(franchise_mock, f)
        
    with open(os.path.join(TEST_SCHEMAS_DIR, "misc_schema.json"), "w", encoding="utf-8") as f:
        json.dump(misc_mock, f)
        
    yield
    
    # Clean up test directories
    if os.path.exists(TEST_SCHEMAS_DIR):
        shutil.rmtree(TEST_SCHEMAS_DIR, ignore_errors=True)
    if os.path.exists(TEST_DATASETS_DIR):
        shutil.rmtree(TEST_DATASETS_DIR, ignore_errors=True)

def test_schema_loader():
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    
    # Matches 'franchise page' to 'franchise_schema.json'
    schema = loader.load_schema("Franchise Page")
    assert schema["dataset_name"] == "test_franchise.xlsx"
    assert schema["sheet_name"] == "Franchise Data"
    
    # Falls back to misc_schema.json for unknown type
    fallback_schema = loader.load_schema("Something completely unknown")
    assert fallback_schema["dataset_name"] == "test_misc.xlsx"

def test_record_normalizer_formatting():
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")
    normalizer = RecordNormalizer(schema)
    
    # Phone number cleaning
    assert normalizer.normalize_value("phone", "+1 (555) 123-4567 ext 90") == "+1 (555) 123-4567 ext 90"
    assert normalizer.normalize_value("phone", "+91-98765#43210") == "+91-9876543210"
    
    # Email cleaning (not in franchise columns, but check sanitization)
    assert normalizer.normalize_value("email", " Contact@AcmeCorp.com ") == "contact@acmecorp.com"
    
    # URL cleaning
    assert normalizer.normalize_value("website", "www.cultfit.com") == "https://www.cultfit.com"
    assert normalizer.normalize_value("website", "http://cult.fit") == "http://cult.fit"
    
    # List formatting joins with |
    assert normalizer.normalize_value("services", ["Site selection", "Marketing"]) == "Site selection | Marketing"
    
    # Whitespace cleanup
    assert normalizer.normalize_value("brand", " Cult    Fit  \n  Gym  ") == "Cult Fit Gym"

def test_record_normalizer_mapping():
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")
    normalizer = RecordNormalizer(schema)
    
    raw_data = {
        "franchise_name": "Cult Gym",
        "investment": "$50,000 - $100,000",
        "phone": "+1 555 123 4567",
        "unmapped_prop": "Extra content value"
    }
    
    normalized = normalizer.normalize_record(
        raw_record=raw_data,
        source_url="https://cultfit.com",
        page_title="Cult Fit Gym Opportunities"
    )
    
    # Maps aliases correctly
    assert normalized["Franchise Name"] == "Cult Gym"
    assert normalized["Investment Required"] == "$50,000 - $100,000"
    assert normalized["Phone"] == "+1 555 123 4567"
    assert normalized["Source URL"] == "https://cultfit.com"
    
    # Places extra properties in Additional Information column as JSON
    assert "unmapped_prop" in normalized["Additional Information"]
    add_info = json.loads(normalized["Additional Information"])
    assert add_info["unmapped_prop"] == "Extra content value"

def test_duplicate_detector():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    
    headers = ["Source URL", "Franchise Name", "Brand", "Last Updated"]
    for idx, col in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=col)
        
    # Append a row
    ws.cell(row=2, column=1, value="https://franchise.com/cult")
    ws.cell(row=2, column=2, value="Cult Gym")
    ws.cell(row=2, column=3, value="Cult")
    
    primary_keys = ["Source URL", "Franchise Name"]
    
    # Duplicate case
    cand_dup = {"Source URL": "https://franchise.com/cult", "Franchise Name": "Cult Gym", "Brand": "Updated Cult"}
    row_idx = DuplicateDetector.find_duplicate_row_index(ws, primary_keys, cand_dup)
    assert row_idx == 2
    
    # New row case
    cand_new = {"Source URL": "https://franchise.com/cult", "Franchise Name": "Anytime Fitness"}
    row_idx_new = DuplicateDetector.find_duplicate_row_index(ws, primary_keys, cand_new)
    assert row_idx_new is None

def test_workbook_manager_insert_and_update():
    manager = WorkbookManager(datasets_dir=TEST_DATASETS_DIR)
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")
    
    records = [
        {
            "Source URL": "https://cultfit.com",
            "Franchise Name": "Cult Center A",
            "Brand": "Cult Fit",
            "Investment Required": "$25k",
            "Phone": "1234",
            "Website": "https://cult.fit",
            "Additional Information": ""
        }
    ]
    
    summary_info = {
        "status": "Success",
        "page_type": "Franchise Listing",
        "strategy": "DIRECT",
        "execution_time": 1200,
        "chunk_count": 1,
        "batch_count": 1,
        "error": ""
    }
    
    # Insert new record
    manager.write_extracted_records(schema, records, "https://cultfit.com", summary_info)
    
    file_path = os.path.join(TEST_DATASETS_DIR, "test_franchise.xlsx")
    assert os.path.exists(file_path)
    
    wb = openpyxl.load_workbook(file_path)
    assert "Franchise Data" in wb.sheetnames
    assert "Extraction Summary" in wb.sheetnames
    
    ws = wb["Franchise Data"]
    assert ws.max_row == 2 # 1 header + 1 record
    assert ws.cell(row=2, column=2).value == "Cult Center A"
    
    # Save first write's dates
    orig_date = ws.cell(row=2, column=8).value # Extraction Date
    
    # Update record
    updated_records = [
        {
            "Source URL": "https://cultfit.com",
            "Franchise Name": "Cult Center A",
            "Brand": "Cult Fit Updated",
            "Investment Required": "$30k",
            "Phone": "1234",
            "Website": "https://cult.fit",
            "Additional Information": ""
        }
    ]
    manager.write_extracted_records(schema, updated_records, "https://cultfit.com", summary_info)
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Franchise Data"]
    assert ws.max_row == 2 # Count shouldn't increase, update in-place
    assert ws.cell(row=2, column=3).value == "Cult Fit Updated"
    assert ws.cell(row=2, column=4).value == "$30k"
    assert ws.cell(row=2, column=8).value == orig_date # Preserved Extraction Date

def test_workbook_manager_failed_url_log():
    manager = WorkbookManager(datasets_dir=TEST_DATASETS_DIR)
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")
    
    manager.log_failed_url(schema, "https://invalidurl.com", "Timeout loading page", retry_recommended=True)
    
    file_path = os.path.join(TEST_DATASETS_DIR, "test_franchise.xlsx")
    wb = openpyxl.load_workbook(file_path)
    assert "Failed URLs" in wb.sheetnames
    
    ws = wb["Failed URLs"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "https://invalidurl.com"
    assert ws.cell(row=2, column=2).value == "Timeout loading page"

@patch("modules.dataset_builder.builder.fetch_webpage")
@patch("modules.dataset_builder.builder.extract_web_data")
@pytest.mark.asyncio
async def test_dataset_builder_e2e(mock_extract, mock_fetch):
    # Mock render results
    mock_fetch.return_value = {
        "url": "https://testfranchise.com",
        "final_url": "https://testfranchise.com",
        "title": "Cult Gym Franchise Opportunity",
        "html": "<html><body>Cult Gym</body></html>",
        "render_time_ms": 500
    }
    
    # Mock extract results
    mock_extract.return_value = ExtractionResult(
        page_title="Cult Gym Franchise Opportunity",
        page_type="Franchise Listing",
        page_summary="Franchise details of Cult Fit.",
        entities=[
            ExtractedEntity(
                entity_type="Franchise Details",
                records=[
                    ExtractedRecord(
                        attributes=[
                            KeyValue(key="franchise_name", value="Cult Gym"),
                            KeyValue(key="investment", value="$60,000"),
                            KeyValue(key="phone", value="+1 222 333 4444")
                        ]
                    )
                ]
            )
        ]
    )
    
    builder = DatasetBuilder(schemas_dir=TEST_SCHEMAS_DIR, datasets_dir=TEST_DATASETS_DIR)
    results = await builder.process_urls("https://testfranchise.com")
    
    assert results["processed"] == 1
    assert results["success"] == 1
    assert results["failed"] == 0
    assert results["details"][0]["status"] == "Success"
    
    # Check that file was created and contains extracted data
    file_path = os.path.join(TEST_DATASETS_DIR, "test_franchise.xlsx")
    assert os.path.exists(file_path)
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Franchise Data"]
    assert ws.max_row in (2, 3) # 1 header + optional previous test record + 1 new record
    # Find matching row for testfranchise
    found = False
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "https://testfranchise.com":
            assert ws.cell(row=r, column=2).value == "Cult Gym"
            assert ws.cell(row=r, column=4).value == "$60,000"
            found = True
    assert found


# ─── RecordMapper unit tests ──────────────────────────────────────────────────

from modules.dataset_builder.schema_mapper import MappingResult
from modules.dataset_builder.record_mapper import (
    RecordMapper,
    PrimaryEntityDetector,
    FieldMapper,
    FieldNormalizer,
    ValidationError,
)


# ── FieldNormalizer ───────────────────────────────────────────────────────────

def test_field_normalizer():
    """FieldNormalizer cleans phones, emails, URLs, currencies, and lists."""
    # Phone
    assert FieldNormalizer.normalize("phone", "+1 (555) 123-4567 ext 90") == "+1 (555) 123-4567 ext 90"
    assert FieldNormalizer.normalize("mobile", "+91-98765#43210") == "+91-9876543210"

    # Email
    assert FieldNormalizer.normalize("email", " Admin@ACME.com ") == "admin@acme.com"

    # URL
    assert FieldNormalizer.normalize("website", "www.cult.fit") == "https://www.cult.fit"
    assert FieldNormalizer.normalize("url", "https://already.ok") == "https://already.ok"

    # List deduplication + pipe join
    result = FieldNormalizer.normalize("services", ["Site selection", "Marketing", "Site selection"])
    assert result == "Site selection | Marketing"

    # Whitespace collapse
    assert FieldNormalizer.normalize("brand", "  Cult   Fit  Gym  ") == "Cult Fit Gym"

    # None → empty string
    assert FieldNormalizer.normalize("anything", None) == ""


# ── FieldMapper ───────────────────────────────────────────────────────────────

def test_field_mapper_aliases():
    """FieldMapper resolves aliases to canonical schema column names."""
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")
    mapper = FieldMapper(schema)

    assert mapper.get_column("investment") == "Investment Required"
    assert mapper.get_column("franchise_name") == "Franchise Name"
    assert mapper.get_column("source_url") == "Source URL"
    assert mapper.get_column("phone") == "Phone"
    assert mapper.get_column("website") == "Website"
    # Unknown key → None
    assert mapper.get_column("totally_unknown_xyz") is None


def test_field_mapper_unmapped():
    """FieldMapper routes unknown keys to the unmapped bucket."""
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")
    mapper = FieldMapper(schema)

    raw = {
        "franchise_name":     "Cult Gym",
        "investment":         "$50,000",
        "mystery_field_xyz":  "unknown value",
    }
    mapped, unmapped = mapper.map(raw)

    assert "Franchise Name" in mapped
    assert "Investment Required" in mapped
    assert "mystery_field_xyz" in unmapped
    assert "Franchise Name" not in unmapped

    # First-non-empty-wins: duplicate alias resolution
    raw2 = {"franchise_name": "Alpha", "franchise name": "Beta"}
    mapped2, _ = mapper.map(raw2)
    assert mapped2["Franchise Name"] == "Alpha"  # first key wins


# ── PrimaryEntityDetector ─────────────────────────────────────────────────────

def test_primary_entity_detection_single():
    """With one entity it is always the primary; related list is empty."""
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")
    detector = PrimaryEntityDetector(schema)

    entities = [
        ExtractedEntity(
            entity_type="Franchise Details",
            records=[ExtractedRecord(attributes=[
                KeyValue(key="franchise_name", value="BeatBox Gym"),
            ])]
        )
    ]
    primary, related = detector.detect(entities)
    assert primary is entities[0]
    assert related == []


def test_primary_entity_detection_multi():
    """Keyword match + single-record bonus selects the correct primary."""
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")
    detector = PrimaryEntityDetector(schema)

    main_entity = ExtractedEntity(
        entity_type="Franchise Details",   # contains "franchise" → +10
        records=[ExtractedRecord(attributes=[
            KeyValue(key="franchise_name", value="BeatBox Gym"),
            KeyValue(key="investment",     value="$50,000"),
        ])]
    )
    related_entity = ExtractedEntity(
        entity_type="Related Opportunities",   # no keyword match
        records=[
            ExtractedRecord(attributes=[KeyValue(key="franchise_name", value="Powerhouse Gym")]),
            ExtractedRecord(attributes=[KeyValue(key="franchise_name", value="My Home Fitness")]),
            ExtractedRecord(attributes=[KeyValue(key="franchise_name", value="Yasmins Body Image")]),
        ]
    )

    primary, related = detector.detect([main_entity, related_entity])
    assert primary is main_entity
    assert len(related) == 1
    assert related[0] is related_entity


# ── RecordMapper ──────────────────────────────────────────────────────────────

def test_record_mapper_franchise():
    """Single franchise entity → exactly one dict with correct column values."""
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")

    result = ExtractionResult(
        page_title="BeatBox Gym Franchise",
        page_type="Franchise Listing",
        page_summary="Fitness franchise opportunity.",
        entities=[
            ExtractedEntity(
                entity_type="Franchise Details",
                records=[ExtractedRecord(attributes=[
                    KeyValue(key="franchise_name", value="BeatBox Gym"),
                    KeyValue(key="investment",     value="$45,000"),
                    KeyValue(key="phone",          value="+1 800 123 4567"),
                    KeyValue(key="website",        value="beatboxgym.com"),
                ])]
            )
        ]
    )

    record = RecordMapper(schema).map(result, "https://beatboxgym.com")

    assert isinstance(record, MappingResult)
    assert record["Franchise Name"] == "BeatBox Gym"
    assert record["Investment Required"] == "$45,000"
    assert record["Source URL"] == "https://beatboxgym.com"
    assert record["Phone"] == "+1 800 123 4567"
    assert record["Website"] == "https://beatboxgym.com"   # prefix added by FieldNormalizer


def test_record_mapper_related_entities_in_additional_info():
    """Related entities are embedded in Additional Information — not extra rows."""
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")

    result = ExtractionResult(
        page_title="Multi-Franchise Page",
        page_type="Franchise Listing",
        page_summary="Multiple franchise opportunities listed.",
        entities=[
            ExtractedEntity(
                entity_type="Franchise Details",
                records=[ExtractedRecord(attributes=[
                    KeyValue(key="franchise_name", value="BeatBox Gym"),
                ])]
            ),
            ExtractedEntity(
                entity_type="Related Opportunities",
                records=[
                    ExtractedRecord(attributes=[KeyValue(key="franchise_name", value="Powerhouse Gym")]),
                    ExtractedRecord(attributes=[KeyValue(key="franchise_name", value="My Home Fitness")]),
                ]
            ),
        ]
    )

    record = RecordMapper(schema).map(result, "https://multi.com")

    # Still ONE record dict
    assert isinstance(record, MappingResult)
    assert record["Franchise Name"] == "BeatBox Gym"

    # Related entities must be inside Additional Information
    assert record["Additional Information"], "Additional Information must not be empty"
    add_info = json.loads(record["Additional Information"])
    assert "related_entities" in add_info

    related = add_info["related_entities"]
    assert len(related) == 1
    assert related[0]["entity_type"] == "Related Opportunities"
    assert len(related[0]["records"]) == 2


def test_record_mapper_validation_error():
    """ValidationError is raised when an unresolvable required field is absent."""
    custom_schema = {
        "dataset_name":    "custom.xlsx",
        "sheet_name":      "Custom",
        "primary_key":     ["Source URL"],
        "required_fields": ["Source URL", "Revenue"],   # "Revenue" cannot be auto-defaulted
        "aliases":         {},
        "columns":         ["Source URL", "Revenue", "Additional Information", "Extraction Date", "Last Updated"],
    }

    result = ExtractionResult(
        page_title="",
        page_type="Unknown",
        page_summary="",
        entities=[]
    )

    with pytest.raises(ValidationError):
        RecordMapper(custom_schema).map(result, "https://test.com")


def test_record_mapper_defaults_for_missing_fields():
    """Required fields get sensible defaults even with no entities."""
    loader = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema = loader.load_schema("Franchise Page")

    result = ExtractionResult(
        page_title="Fallback Gym",
        page_type="Franchise Listing",
        page_summary="",
        entities=[]   # no entities at all
    )

    record = RecordMapper(schema).map(result, "https://fallback.com")

    # Required fields are filled by defaults
    assert record["Source URL"] == "https://fallback.com"
    # "Franchise Name" defaults to page_title
    assert record["Franchise Name"] == "Fallback Gym"


def test_record_mapper_one_row_per_url():
    """Integration: RecordMapper + WorkbookManager always produces exactly ONE new Excel row."""
    loader  = SchemaLoader(schemas_dir=TEST_SCHEMAS_DIR)
    schema  = loader.load_schema("Franchise Page")
    manager = WorkbookManager(datasets_dir=TEST_DATASETS_DIR)

    result = ExtractionResult(
        page_title="Alpha Gym Franchise",
        page_type="Franchise Listing",
        page_summary="Alpha Gym opportunity.",
        entities=[
            ExtractedEntity(
                entity_type="Franchise Details",
                records=[ExtractedRecord(attributes=[
                    KeyValue(key="franchise_name", value="Alpha Gym"),
                ])]
            ),
            # Multi-record related entity — must NOT create extra rows
            ExtractedEntity(
                entity_type="Related Franchises",
                records=[
                    ExtractedRecord(attributes=[KeyValue(key="franchise_name", value="Beta Gym")]),
                    ExtractedRecord(attributes=[KeyValue(key="franchise_name", value="Gamma Gym")]),
                ]
            ),
        ]
    )

    record = RecordMapper(schema).map(result, "https://alphagym.com")

    rows_before = _count_data_rows(TEST_DATASETS_DIR, schema)

    summary_info = {
        "status": "Success", "page_type": "Franchise Listing",
        "strategy": "DIRECT", "execution_time": 0,
        "chunk_count": 1, "batch_count": 1, "error": "",
    }
    manager.write_extracted_records(schema, [record], "https://alphagym.com", summary_info)

    rows_after = _count_data_rows(TEST_DATASETS_DIR, schema)
    # Exactly ONE row added regardless of how many entities/records existed
    assert rows_after == rows_before + 1


def _count_data_rows(datasets_dir: str, schema: dict) -> int:
    """Helper: data rows in the schema's sheet (header excluded)."""
    file_path = os.path.join(datasets_dir, schema["dataset_name"])
    if not os.path.exists(file_path):
        return 0
    wb = openpyxl.load_workbook(file_path)
    sheet_name = schema["sheet_name"]
    if sheet_name not in wb.sheetnames:
        return 0
    ws = wb[sheet_name]
    return max(0, ws.max_row - 1)


def test_save_extraction_result_direct():
    result = ExtractionResult(
        page_title="Cult Gym Franchise Opportunity",
        page_type="Franchise Listing",
        page_summary="Franchise details of Cult Fit.",
        entities=[
            ExtractedEntity(
                entity_type="Franchise Details",
                records=[
                    ExtractedRecord(
                        attributes=[
                            KeyValue(key="franchise_name", value="Cult Gym Direct"),
                            KeyValue(key="investment", value="$70,000"),
                            KeyValue(key="phone", value="+1 222 333 4444")
                        ]
                    )
                ]
            )
        ]
    )
    builder = DatasetBuilder(schemas_dir=TEST_SCHEMAS_DIR, datasets_dir=TEST_DATASETS_DIR)
    
    # Save first time (should be Inserted)
    res_save = builder.save_extraction_result(
        result=result,
        source_url="https://testfranchisedirect.com",
        detected_page_type="Franchise Listing",
        timestamp="2026-07-02 12:00:00"
    )
    assert res_save["status"] == "Success"
    assert res_save["operation"] == "Inserted"
    
    # Save second time (should be Updated)
    res_save_up = builder.save_extraction_result(
        result=result,
        source_url="https://testfranchisedirect.com",
        detected_page_type="Franchise Listing",
        timestamp="2026-07-02 12:05:00"
    )
    assert res_save_up["status"] == "Success"
    assert res_save_up["operation"] == "Updated"
    assert res_save_up["row_number"] == res_save["row_number"]

def test_locked_excel_file_handling():
    import asyncio
    from modules.dataset_builder.manager import ExcelFileLockedError
    
    result = ExtractionResult(
        franchise_name="Locked Gym",
        page_title="Locked Gym Opportunity",
        page_type="Franchise Listing"
    )
    builder = DatasetBuilder(schemas_dir=TEST_SCHEMAS_DIR, datasets_dir=TEST_DATASETS_DIR)
    
    # Mock ExcelWriter.save_workbook to raise PermissionError simulating locked file
    with patch("modules.dataset_builder.manager.openpyxl.Workbook.save", side_effect=PermissionError("Permission denied")):
        res_save = builder.save_extraction_result(
            result=result,
            source_url="https://testlocked.com",
            detected_page_type="Franchise Listing",
            timestamp="2026-07-02 12:00:00"
        )
        assert res_save["status"] == "Failed"
        assert "locked/open" in res_save["reason"]
        assert "File Locked" in res_save["operation"]

        # Run process_urls and verify it maps locked failure without crashing
        with patch("modules.dataset_builder.builder.fetch_webpage", return_value={"html": "<html></html>", "title": "Locked"}):
            with patch("modules.dataset_builder.builder.extract_web_data", return_value=result):
                with patch("modules.dataset_builder.builder.WorkbookManager.write_extracted_records", side_effect=ExcelFileLockedError("Locked file error")):
                    batch_res = asyncio.run(builder.process_urls(["https://testlocked.com"]))
                    assert batch_res["failed"] == 1
                    assert batch_res["details"][0]["status"] == "Failed (File Locked)"



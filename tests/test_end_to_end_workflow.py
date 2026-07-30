"""
Full-system integration test, exercising the whole stack as a caller (a
future UI, a script, another service) would: never touching a store or the
pipeline directly, only the three services.

    Create Config -> Save -> Load
    Create Schema -> Save -> Load
    Create Job -> Run -> Output row written

Note: the pipeline's final stage writes via DatasetBuilder, which is
format-agnostic (writers/excel_writer.py or writers/csv_writer.py, chosen
via output_format - see writers/dataset_writer.py). This test verifies the
default output_format ("excel"), an .xlsx row; tests/test_dataset_writer.py
covers the CSV path and the Excel/CSV column-order regression check.
"""

import os
import tempfile
from unittest.mock import patch

import openpyxl
import pytest

from config.config_store import ConfigStore
from config.schema_store import SchemaStore
from config.job_store import JobStore
from services.config_service import ConfigService
from services.schema_service import SchemaService
from services.job_service import JobService


HTML = """
<html><head><title>Cult Fit Franchise Opportunity</title></head>
<body>
  <h1>Cult Fit Franchise Opportunity</h1>
  <table>
    <tr><td>Investment:</td><td>Rs. 30 Lakhs - 50 Lakhs</td></tr>
  </table>
  <a href="mailto:contact@cultfit.com">Email</a>
</body></html>
"""


async def _fake_fetch_webpage(url, timeout_ms=30000):
    return {"url": url, "final_url": url, "title": "Cult Fit", "html": HTML, "render_time_ms": 1}


def _llm_side_effect(**kwargs):
    model_cls = kwargs["response_model"]
    return model_cls(page_title="Cult Fit", page_summary="A fitness franchise", business_model="Gym franchise model")


@pytest.mark.asyncio
async def test_full_workflow_create_config_schema_job_run_writes_output(tmp_path):
    config_service = ConfigService(ConfigStore(storage_dir=str(tmp_path / "configs")))
    schema_service = SchemaService(SchemaStore(storage_dir=str(tmp_path / "schemas")))

    # 1. Create Config -> Save
    config_id = config_service.create({
        "name": "Generic Franchise Directory",
        "domain": "example-directory.com",
        "browser_config": {"wait_strategy": {"wait_until": "domcontentloaded"}},
    })

    # 2. Load it back (verifies it actually persisted, not just held in memory)
    reloaded_config = config_service.get(config_id)
    assert reloaded_config.name == "Generic Franchise Directory"

    # 3. Create Schema -> Save -> Load
    schema_id = schema_service.create({
        "name": "Franchise Listing",
        "fields": [
            {"name": "franchise_name", "type": "string", "description": "Name of the franchise"},
            {"name": "investment_required", "type": "string", "description": "Investment required"},
            {"name": "business_model", "type": "string", "description": "Business model summary"},
        ],
    })
    reloaded_schema = schema_service.get(schema_id)
    assert reloaded_schema.field_names() == ["franchise_name", "investment_required", "business_model"]

    with tempfile.TemporaryDirectory() as tmp_datasets:
        job_service = JobService(
            config_store=config_service.store,
            schema_store=schema_service.store,
            job_store=JobStore(storage_dir=str(tmp_path / "jobs")),
        )

        # 4. Create Job
        job = job_service.create_job(
            name="Franchise Directory Extraction",
            urls=["https://example-directory.com/franchise/cult-fit"],
            config_id=config_id,
            schema_id=schema_id,
        )
        assert job.status == "created"

        # 5. Run
        with patch("modules.dataset_builder.builder.fetch_webpage", new=_fake_fetch_webpage), \
             patch("modules.dataset_builder.builder.extract_web_data") as mock_extract, \
             patch.dict(os.environ, {"GEMINI_API_KEY": "test"}):
            mock_extract.side_effect = _llm_side_effect

            completed_job = await job_service.run(job.id, datasets_dir=tmp_datasets)

        assert completed_job.status == "completed"
        assert len(completed_job.stage_log) == 1
        assert completed_job.stage_log[0]["status"] == "success"

        # A fresh load from the JobStore reflects the same final state.
        reloaded_job = job_service.get(job.id)
        assert reloaded_job.status == "completed"

        # 6. Output row actually written (default output_format is "excel";
        # tests/test_dataset_writer.py exercises the CSV path).
        save_info = completed_job.stage_log[0]["save_info"]
        workbook_path = os.path.join(tmp_datasets, os.path.basename(save_info["workbook_name"]))
        assert os.path.exists(workbook_path)

        wb = openpyxl.load_workbook(workbook_path)
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in ws[1]]
        row = [cell.value for cell in ws[2]]
        row_dict = dict(zip(headers, row))

        assert row_dict.get("Franchise Name") == "Cult Fit Franchise Opportunity"
        assert "30 Lakhs" in (row_dict.get("Investment Required") or "")
        assert row_dict.get("Business Model") == "Gym franchise model"

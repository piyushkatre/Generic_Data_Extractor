from typing import List, Dict, Optional
from openpyxl.worksheet.worksheet import Worksheet

class DuplicateDetector:
    """
    Scans a worksheet to locate rows matching primary key constraints.
    """

    @staticmethod
    def find_duplicate_row_index(
        ws: Worksheet,
        primary_keys: List[str],
        candidate_record: Dict[str, str]
    ) -> Optional[int]:
        """
        Scans all rows of the worksheet to check if any row matches the candidate_record
        on all specified primary key columns.
        Returns the 1-based row index if a match is found, otherwise None.
        """
        if ws.max_row < 2:
            return None

        # Build column index mapping (header on row 1)
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        col_indices = {}
        for pk in primary_keys:
            try:
                # openpyxl columns are 1-indexed
                idx = headers.index(pk) + 1
                col_indices[pk] = idx
            except ValueError:
                # Primary key column not found in headers
                continue

        if not col_indices:
            return None

        # Check each row starting from row 2
        for r in range(2, ws.max_row + 1):
            match = True
            for pk, col_idx in col_indices.items():
                cell_val = ws.cell(row=r, column=col_idx).value
                candidate_val = candidate_record.get(pk, "")
                
                # Normalize values for clean comparison
                str_cell = str(cell_val if cell_val is not None else "").strip().lower()
                str_cand = str(candidate_val if candidate_val is not None else "").strip().lower()
                
                if str_cell != str_cand:
                    match = False
                    break
            if match:
                return r

        return None

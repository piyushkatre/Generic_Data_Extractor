"""
Excel (.xlsx) dataset output.

Everything below `ExcelDatasetWriter` (ExcelFileLockedError, ExcelWriter,
WorkbookManager, DuplicateDetector) is moved here verbatim from
modules/dataset_builder/manager.py and modules/dataset_builder/detector.py -
no behavior changes, just relocated so all Excel-specific/openpyxl-specific
code lives in one place. Those two original modules are now thin re-export
shims over this file, kept for backward compatibility with existing
callers/tests that import from the old locations.

`ExcelDatasetWriter` is the new part: a thin adapter implementing the
format-agnostic DatasetWriter interface (writers/dataset_writer.py) by
driving the same WorkbookManager/ExcelWriter/DuplicateDetector logic that
was previously inlined a second time inside DatasetBuilder.save_extraction_result().
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from writers.dataset_writer import DatasetWriter


class ExcelFileLockedError(PermissionError):
    """Exception raised when an Excel workbook file is locked or open in another application."""
    pass


class ExcelWriter:
    """
    Directly interacts with openpyxl to perform workbook file I/O operations.
    """

    @staticmethod
    def get_or_create_workbook(file_path: str) -> openpyxl.Workbook:
        if os.path.exists(file_path):
            try:
                # Exclusive lock check to early-detect locked file
                with open(file_path, "r+"):
                    pass
            except PermissionError as e:
                raise ExcelFileLockedError(
                    f"The Excel file '{os.path.basename(file_path)}' is currently locked/open. Please close it and retry."
                ) from e
            except Exception:
                pass

            try:
                return openpyxl.load_workbook(file_path)
            except Exception:
                # If file is corrupted or empty, create a new one
                pass
        return openpyxl.Workbook()

    @staticmethod
    def save_workbook(wb: openpyxl.Workbook, file_path: str):
        # Ensure parent directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        try:
            wb.save(file_path)
        except PermissionError as e:
            raise ExcelFileLockedError(
                f"The Excel file '{os.path.basename(file_path)}' is currently locked/open. Please close it and retry."
            ) from e


class WorkbookManager:
    """
    Orchestrates sheet structure, records insertion/update, extraction summaries,
    and failed extraction logs inside openpyxl Workbooks.
    """

    def __init__(self, datasets_dir: str = "datasets"):
        self.datasets_dir = os.path.abspath(datasets_dir)

    def write_extracted_records(
        self,
        schema: Dict[str, Any],
        normalized_records: List[Dict[str, str]],
        source_url: str,
        summary_info: Dict[str, Any]
    ):
        """
        Appends or updates extracted data records in the target dataset workbook,
        and logs the operation in the Extraction Summary sheet.
        """
        dataset_name = schema.get("dataset_name", "misc_dataset.xlsx")
        sheet_name = schema.get("sheet_name", "General Web Data")
        primary_keys = schema.get("primary_key", [])
        columns = schema.get("columns", [])

        file_path = os.path.join(self.datasets_dir, dataset_name)
        wb = ExcelWriter.get_or_create_workbook(file_path)

        # Ensure sheets exist
        self._ensure_sheet_headers(wb, sheet_name, columns)
        self._ensure_summary_sheet_headers(wb)

        ws = wb[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

        timestamp_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for record in normalized_records:
            # Check for existing record matching primary keys
            row_idx = DuplicateDetector.find_duplicate_row_index(ws, primary_keys, record)

            # Map column names to index values (1-indexed for openpyxl)
            if row_idx:
                # Update existing row
                # Set Last Updated, preserve Extraction Date if present
                for col in columns:
                    if col == "Last Updated":
                        record[col] = timestamp_now
                    elif col == "Extraction Date":
                        # Fetch original Extraction Date value
                        try:
                            col_idx = headers.index(col) + 1
                            orig_val = ws.cell(row=row_idx, column=col_idx).value
                            record[col] = str(orig_val if orig_val is not None else timestamp_now)
                        except ValueError:
                            record[col] = timestamp_now

                    if col in record:
                        try:
                            idx = headers.index(col) + 1
                            ws.cell(row=row_idx, column=idx, value=record[col])
                        except ValueError:
                            pass
            else:
                # Insert new row
                # Set both Extraction Date and Last Updated
                row_idx = ws.max_row + 1
                for col in columns:
                    if col == "Extraction Date" or col == "Last Updated":
                        record[col] = timestamp_now

                    if col in record:
                        try:
                            idx = headers.index(col) + 1
                            ws.cell(row=row_idx, column=idx, value=record[col])
                        except ValueError:
                            pass

        # Log summary info
        self._log_summary(wb, source_url, summary_info)
        ExcelWriter.save_workbook(wb, file_path)

        return {
            "operation": "Updated" if row_idx else "Inserted",
            "row_number": row_idx if row_idx else ws.max_row
        }

    def log_failed_url(
        self,
        schema: Dict[str, Any],
        url: str,
        reason: str,
        retry_recommended: bool = True
    ):
        """
        Logs a failed URL extraction into the Failed URLs sheet of the workbook.
        """
        dataset_name = schema.get("dataset_name", "misc_dataset.xlsx")
        file_path = os.path.join(self.datasets_dir, dataset_name)
        wb = ExcelWriter.get_or_create_workbook(file_path)

        self._ensure_failed_sheet_headers(wb)
        ws = wb["Failed URLs"]

        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=url)
        ws.cell(row=row_idx, column=2, value=reason)
        ws.cell(row=row_idx, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_idx, column=4, value="Yes" if retry_recommended else "No")

        ExcelWriter.save_workbook(wb, file_path)

    def _ensure_sheet_headers(self, wb: openpyxl.Workbook, sheet_name: str, columns: List[str]):
        """
        Checks if the target data sheet exists and has correct headers. If not, initializes it.
        """
        if sheet_name not in wb.sheetnames:
            # If sheet is new and there's a default Sheet created, rename it
            if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
                ws = wb["Sheet"]
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            # Write headers
            for idx, col in enumerate(columns, start=1):
                ws.cell(row=1, column=idx, value=col)
        else:
            # Sheet exists, verify headers match
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            if not headers or all(h is None for h in headers):
                for idx, col in enumerate(columns, start=1):
                    ws.cell(row=1, column=idx, value=col)

    def _ensure_summary_sheet_headers(self, wb: openpyxl.Workbook):
        """
        Initializes the 'Extraction Summary' sheet.
        """
        summary_cols = [
            "URL", "Status", "Page Type", "Strategy", "Execution Time (ms)",
            "Chunk Count", "Batch Count", "Extraction Timestamp", "Error"
        ]
        sheet_name = "Extraction Summary"
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            for idx, col in enumerate(summary_cols, start=1):
                ws.cell(row=1, column=idx, value=col)
        else:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            if not headers or all(h is None for h in headers):
                for idx, col in enumerate(summary_cols, start=1):
                    ws.cell(row=1, column=idx, value=col)

    def _ensure_failed_sheet_headers(self, wb: openpyxl.Workbook):
        """
        Initializes the 'Failed URLs' sheet.
        """
        failed_cols = ["URL", "Reason", "Timestamp", "Retry Recommended"]
        sheet_name = "Failed URLs"
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            for idx, col in enumerate(failed_cols, start=1):
                ws.cell(row=1, column=idx, value=col)
        else:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            if not headers or all(h is None for h in headers):
                for idx, col in enumerate(failed_cols, start=1):
                    ws.cell(row=1, column=idx, value=col)

    def _log_summary(self, wb: openpyxl.Workbook, url: str, info: Dict[str, Any]):
        """
        Appends an extraction summary record.
        """
        ws = wb["Extraction Summary"]
        row_idx = ws.max_row + 1

        ws.cell(row=row_idx, column=1, value=url)
        ws.cell(row=row_idx, column=2, value=info.get("status", "Success"))
        ws.cell(row=row_idx, column=3, value=info.get("page_type", "Unknown"))
        ws.cell(row=row_idx, column=4, value=info.get("strategy", "DIRECT"))
        ws.cell(row=row_idx, column=5, value=info.get("execution_time", 0))
        ws.cell(row=row_idx, column=6, value=info.get("chunk_count", 1))
        ws.cell(row=row_idx, column=7, value=info.get("batch_count", 1))
        ws.cell(row=row_idx, column=8, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_idx, column=9, value=info.get("error", ""))


class DuplicateDetector:
    """
    Scans a worksheet to locate rows matching primary key constraints.
    """

    @staticmethod
    def find_duplicate_row_index(
        ws: Worksheet,
        primary_keys: List[str],
        candidate_record: Dict[str, str]
    ) -> Optional[int]:
        """
        Scans all rows of the worksheet to check if any row matches the candidate_record
        on all specified primary key columns.
        Returns the 1-based row index if a match is found, otherwise None.
        """
        if ws.max_row < 2:
            return None

        # Build column index mapping (header on row 1)
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        col_indices = {}
        for pk in primary_keys:
            try:
                # openpyxl columns are 1-indexed
                idx = headers.index(pk) + 1
                col_indices[pk] = idx
            except ValueError:
                # Primary key column not found in headers
                continue

        if not col_indices:
            return None

        # Check each row starting from row 2
        for r in range(2, ws.max_row + 1):
            match = True
            for pk, col_idx in col_indices.items():
                cell_val = ws.cell(row=r, column=col_idx).value
                candidate_val = candidate_record.get(pk, "")

                # Normalize values for clean comparison
                str_cell = str(cell_val if cell_val is not None else "").strip().lower()
                str_cand = str(candidate_val if candidate_val is not None else "").strip().lower()

                if str_cell != str_cand:
                    match = False
                    break
            if match:
                return r

        return None


class ExcelDatasetWriter(DatasetWriter):
    """
    DatasetWriter implementation for .xlsx output. Drives the WorkbookManager/
    ExcelWriter/DuplicateDetector logic above - the same logic that used to be
    duplicated inline inside DatasetBuilder.save_extraction_result() - through
    the format-agnostic begin_dataset/write_headers/append_record/finish
    interface instead.
    """

    def __init__(self, datasets_dir: str = "datasets"):
        self.datasets_dir = os.path.abspath(datasets_dir)
        self._manager = WorkbookManager(datasets_dir=datasets_dir)
        self._file_path: Optional[str] = None
        self._sheet_name: Optional[str] = None
        self._columns: List[str] = []
        self._wb: Optional[openpyxl.Workbook] = None
        self._ws: Optional[Worksheet] = None
        self._headers: List[str] = []

    def begin_dataset(self, dataset_name: str, sheet_name: str, columns: List[str]) -> None:
        dataset_name = dataset_name or "misc_dataset.xlsx"
        base, ext = os.path.splitext(dataset_name)
        if ext.lower() != ".xlsx":
            dataset_name = f"{base}.xlsx"

        self._file_path = os.path.join(self.datasets_dir, dataset_name)
        self._sheet_name = sheet_name or "General Web Data"
        self._columns = list(columns or [])
        self._wb = ExcelWriter.get_or_create_workbook(self._file_path)

    def write_headers(self) -> None:
        self._manager._ensure_sheet_headers(self._wb, self._sheet_name, self._columns)
        self._manager._ensure_summary_sheet_headers(self._wb)
        self._ws = self._wb[self._sheet_name]
        self._headers = [str(cell.value).strip() if cell.value is not None else "" for cell in self._ws[1]]

    def append_record(self, record: Dict[str, Any], primary_keys: List[str], timestamp: str) -> Dict[str, Any]:
        row_idx = DuplicateDetector.find_duplicate_row_index(self._ws, primary_keys, record)

        if row_idx:
            operation = "Updated"
            for col in self._columns:
                if col == "Last Updated":
                    record[col] = timestamp
                elif col == "Extraction Date":
                    try:
                        col_idx = self._headers.index(col) + 1
                        orig_val = self._ws.cell(row=row_idx, column=col_idx).value
                        record[col] = str(orig_val if orig_val is not None else timestamp)
                    except ValueError:
                        record[col] = timestamp

                if col in record:
                    try:
                        idx = self._headers.index(col) + 1
                        self._ws.cell(row=row_idx, column=idx, value=record[col])
                    except ValueError:
                        pass
        else:
            operation = "Inserted"
            row_idx = self._ws.max_row + 1
            for col in self._columns:
                if col == "Extraction Date" or col == "Last Updated":
                    record[col] = timestamp

                if col in record:
                    try:
                        idx = self._headers.index(col) + 1
                        self._ws.cell(row=row_idx, column=idx, value=record[col])
                    except ValueError:
                        pass

        return {"operation": operation, "row_number": row_idx}

    def log_summary(self, source_url: str, summary_info: Dict[str, Any]) -> None:
        self._manager._log_summary(self._wb, source_url, summary_info)

    def log_failed_url(self, url: str, reason: str, retry_recommended: bool = True) -> None:
        dataset_name = os.path.basename(self._file_path) if self._file_path else "misc_dataset.xlsx"
        self._manager.log_failed_url({"dataset_name": dataset_name}, url, reason, retry_recommended)

    def finish(self) -> Dict[str, Any]:
        ExcelWriter.save_workbook(self._wb, self._file_path)
        record_count = max(self._ws.max_row - 1, 0) if self._ws is not None else 0
        return {
            "file_path": self._file_path,
            "format": "excel",
            "record_count": record_count,
            # The workbook also has "Extraction Summary"/"Failed URLs"
            # sheets - a reader needs the actual data sheet's name, not
            # just the file path, to avoid reading the wrong one.
            "sheet_name": self._sheet_name,
        }
